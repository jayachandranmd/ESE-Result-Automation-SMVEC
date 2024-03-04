from flask import Flask, render_template, request, send_file
from werkzeug.utils import secure_filename
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import zipfile

app = Flask(__name__)

UPLOAD_FOLDER = 'uploads'
RESULT_FOLDER = 'results'
ALLOWED_EXTENSIONS = {'xlsx'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['RESULT_FOLDER'] = RESULT_FOLDER

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def process_results(result_url, file_path):
    # Load Excel file
    wb = load_workbook(file_path)
    ws = wb.active

    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--start-maximized')
    driver = webdriver.Chrome(options=options)

    driver.get(result_url)

    result_wb = Workbook()
    result_ws = result_wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= 2:
            regno, dob = row[:2]
            while True:
                try:
                    # Defining the input fields based on ID 
                    resigter_number_field = driver.find_element(By.ID,"txtRollNo")
                    dob_field = driver.find_element(By.ID,"txtDoB")
                    captcha_text = driver.find_element(By.ID,"mainCaptcha")

                    # Extracting captcha text
                    captcha_text_value = captcha_text.text
                    captcha_field = driver.find_element(By.ID,"txtInput")

                    # Input values from excel to result site
                    resigter_number_field.click()
                    resigter_number_field.send_keys(regno)
                    dob_field.click()
                    dob_field.send_keys(dob)
                    captcha_field.click()
                    captcha_field.send_keys(captcha_text_value)
                    captcha_field.send_keys(Keys.ENTER)

                    time.sleep(3)  

                    try:
                        # Extracting name 
                        name_element = driver.find_element(By.XPATH,"//div[@id='printdataResults']/div[5]")
                        name = name_element.text.split(": ")[1] if ' ' in name_element.text else ''

                        # Extracting SGPA
                        sgpa_element = driver.find_element(By.XPATH,"//div[@id='printdataResults']/div[7]")
                        sgpa = sgpa_element.text.split(maxsplit=1)[1] if ' ' in sgpa_element.text else ''

                        # Printing Name and SGPA
                        print(name)
                        print(sgpa)

                        # Finding number of subjects
                        table_element = driver.find_element(By.XPATH,"//div[@id='printdataResults']/div[6]/table/tbody")
                        tr_elements = table_element.find_elements(By.TAG_NAME, "tr")
                        tr_count = len(tr_elements)

                        # Creating empty lists to store the Column Attributes and Subjects 
                        title_list = []
                        subject_list = []

                        # Looping through the HTML tags to obtain the Titles list and Subjects list
                        for i in range(1,tr_count+1): 
                            title_xpath_expression = "//div[@id='printdataResults']/div[6]/table/tbody/tr[{}]/td[3]".format(i)
                            subject_xpath_expression = "//div[@id='printdataResults']/div[6]/table/tbody/tr[{}]/td[5]".format(i)
                            title_i_element = driver.find_element(By.XPATH, title_xpath_expression)
                            sub_i_element = driver.find_element(By.XPATH, subject_xpath_expression)
                            subject_list.append(sub_i_element.text)
                            title_list.append(title_i_element.text)

                        # Ensuring that Column Attributes are set only once 
                        if 'flag' not in locals():

                            # Setting Column Attributes 
                            first_row = ["Register Number","Name"]
                            first_row.extend(title_list)
                            first_row.append("SGPA")
                            result_ws.append(first_row)
                            flag = 1

                        time.sleep(2)

                        print(subject_list)
                        print(title_list)

                        # Appending row wise values 
                        row = [regno,name]
                        row.extend(subject_list)
                        row.append(sgpa)

                        # Appending values to result.xlsx
                        result_ws.append(row)

                    except Exception as e:
                        print("Error:", e)
                        continue
                    break  

                except Exception as e:
                    print("Error:", e)
                    continue

    result_wb.save(os.path.join(app.config['RESULT_FOLDER'], 'result.xlsx'))
    driver.quit()


@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        # Check if the post request has the file part
        if 'file' not in request.files:
            return render_template('index.html', error='No file part')
        
        file = request.files['file']
        # If the user does not select a file, the browser submits an empty file without a filename
        if file.filename == '':
            return render_template('index.html', error='No selected file')
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)

            # Get exam result site URL from form input
            result_url = request.form['result_url']

            process_results(result_url, file_path)
            return send_file(os.path.join(app.config['RESULT_FOLDER'], 'result.xlsx'), as_attachment=True)

    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=False)


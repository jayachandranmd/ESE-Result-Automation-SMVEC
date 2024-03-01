from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import load_workbook
from openpyxl import Workbook
from selenium.webdriver.common.by import By

# Input file containing Register number and DOB of students
# First column must contain Register number and Second Column must contain DOB in the format dd/mm/yyyy as String 
wb = load_workbook(r'C:\Users\jaich\Desktop\student_data.xlsx') 
ws = wb.active

# Defining the browser type 
options = webdriver.ChromeOptions()
options.add_argument('--headless')
options.add_argument('--start-maximized')
driver = webdriver.Chrome(options=options)

# Exam result site
driver.get("https://exam.smvec.ac.in/exam_result_ug_pg_regular_jan_2024/")

# Excel file for stroring the result
result_wb = Workbook()
result_ws = result_wb.active

flag = 0
for row in ws.iter_rows(min_row=11, values_only=True):
    if len(row) >= 2:
        regno, dob = row[:2]
        while True:
            try:
                # Defining the input fields based on ID 
                resigter_number_field = driver.find_element(By.ID,"txtRollNo")
                dob_field = driver.find_element(By.ID,"txtDoB")
                captcha_text = driver.find_element(By.ID,"mainCaptcha")

                # Extracting captcha text
                captcha_text_value = captcha_text.text
                captcha_field = driver.find_element(By.ID,"txtInput")
                
                # Input values from excel to result site
                resigter_number_field.click()
                resigter_number_field.send_keys(regno)
                dob_field.click()
                dob_field.send_keys(dob)
                captcha_field.click()
                captcha_field.send_keys(captcha_text_value)
                captcha_field.send_keys(Keys.ENTER)

                time.sleep(3)  
                
                try:
                    # Extracting name 
                    name_element = driver.find_element(By.XPATH,"//div[@id='printdataResults']/div[5]")
                    name = name_element.text.split(": ")[1] if ' ' in name_element.text else ''

                    # Extracting SGPA
                    sgpa_element = driver.find_element(By.XPATH,"//div[@id='printdataResults']/div[7]")
                    sgpa = sgpa_element.text.split(maxsplit=1)[1] if ' ' in sgpa_element.text else ''

                    # Printing Name and SGPA
                    print(name)
                    print(sgpa)

                    # Finding number of subjects
                    table_element = driver.find_element(By.XPATH,"//div[@id='printdataResults']/div[6]/table/tbody")
                    tr_elements = table_element.find_elements(By.TAG_NAME, "tr")
                    tr_count = len(tr_elements)
                    
                    # Creating empty lists to store the Column Attributes and Subjects 
                    title_list = []
                    subject_list = []

                    # Looping through the HTML tags to obtain the Titles list and Subjects list
                    for i in range(1,tr_count+1): 
                        title_xpath_expression = "//div[@id='printdataResults']/div[6]/table/tbody/tr[{}]/td[3]".format(i)
                        subject_xpath_expression = "//div[@id='printdataResults']/div[6]/table/tbody/tr[{}]/td[5]".format(i)
                        title_i_element = driver.find_element(By.XPATH, title_xpath_expression)
                        sub_i_element = driver.find_element(By.XPATH, subject_xpath_expression)
                        subject_list.append(sub_i_element.text)
                        title_list.append(title_i_element.text)
                    
                    # Ensuring that screen dimensions and Column Attributes are set only once 
                    if flag == 0:
                        # Setting screen dimensions
                        width = driver.execute_script("return document.body.scrollWidth")
                        height = driver.execute_script("return document.body.scrollHeight")
                        driver.set_window_size(width, height) 

                        # Setting Column Attributes 
                        first_row = ["Register Number","Name"]
                        first_row.extend(title_list)
                        first_row.append("SGPA")
                        result_ws.append(first_row)
                        flag = 1
                    
                    time.sleep(2)

                    print(subject_list)
                    print(title_list)

                    # Appending row wise values 
                    row = [regno,name]
                    row.extend(subject_list)
                    row.append(sgpa)

                    # Appending values to result.xlsx
                    result_ws.append(row)

                    # Save screenshot of result with register number as file name
                    driver.save_screenshot(f"result/{regno}_result.png")
                    
                except Exception as e:
                    print("Error:", e)
                    continue
                break  

            except Exception as e:
                print("Error:", e)
                continue

# Saving the Excel file to the project directory 
result_wb.save("result.xlsx")

#Quit the driver
driver.quit()
